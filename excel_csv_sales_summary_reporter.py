import json
import math
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_DIR = "input"
MASTER_DIR = "master"
OUTPUT_DIR = "output"
CONFIG_FILE = "config.json"

SUPPORTED_INPUT_EXTENSIONS = [".csv", ".xlsx"]
EXCEL_EXTENSION = ".xlsx"
CSV_EXTENSION = ".csv"

DEFAULT_CSV_READ_ENCODINGS = ["utf-8-sig", "utf-8", "cp932"]
DEFAULT_DATE_FORMAT = "%Y-%m-%d"
DEFAULT_MONTH_FORMAT = "%Y-%m"

SOURCE_FILE_COLUMN = "source_file"
ROW_NUMBER_COLUMN = "row_number"
INTERNAL_MASTER_ROW_NUMBER = "__master_row_number"
MERGE_STATUS_COLUMN = "_merge"
ERROR_TYPE_COLUMN = "error_type"
ERROR_MESSAGE_COLUMN = "error_message"

DEFAULT_SHEET_NAMES = {
    "summary": "summary",
    "sales_details": "sales_details",
    "monthly_sales": "monthly_sales",
    "category_sales": "category_sales",
    "product_sales": "product_sales",
    "unmatched_products": "unmatched_products",
    "validation_errors": "validation_errors",
}

DEFAULT_SALES_DETAILS_OUTPUT_COLUMNS = [
    "source_file",
    "row_number",
    "sales_id",
    "sales_date",
    "sales_month",
    "product_code",
    "product_name",
    "category",
    "quantity",
    "unit_price",
    "sales_amount",
]

DEFAULT_MONTHLY_SALES_OUTPUT_COLUMNS = [
    "sales_month",
    "sales_count",
    "total_quantity",
    "total_sales",
]

DEFAULT_CATEGORY_SALES_OUTPUT_COLUMNS = [
    "category",
    "sales_count",
    "total_quantity",
    "total_sales",
]

DEFAULT_PRODUCT_SALES_OUTPUT_COLUMNS = [
    "product_code",
    "product_name",
    "category",
    "sales_count",
    "total_quantity",
    "total_sales",
]

DEFAULT_ERROR_OUTPUT_COLUMNS = [
    "source_file",
    "row_number",
    "sales_id",
    "sales_date",
    "product_code",
    "quantity",
    "unit_price",
    "error_type",
    "error_message",
]


def create_sample_files():
    """Create folders, config.json, sample sales files, and sample master data."""
    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(MASTER_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not os.path.exists(CONFIG_FILE):
        sample_config = {
            "sales_files": [
                "sample_sales_2026_06.xlsx",
                "sample_sales_2026_07.xlsx"
            ],
            "master_file": "product_master.xlsx",
            "report_file": "sales_summary_report.xlsx",
            "sales_key": "product_code",
            "master_key": "product_code",
            "sales_required_columns": [
                "sales_id",
                "sales_date",
                "product_code",
                "quantity",
                "unit_price"
            ],
            "master_required_columns": [
                "product_code",
                "product_name",
                "category"
            ],
            "master_value_columns": [
                "product_name",
                "category"
            ],
            "sales_numeric_columns": [
                "quantity",
                "unit_price"
            ],
            "sales_date_columns": [
                "sales_date"
            ],
            "date_format": "%Y-%m-%d",
            "month_format": "%Y-%m",
            "amount_calculation": {
                "enabled": True,
                "quantity_column": "quantity",
                "price_column": "unit_price",
                "output_column": "sales_amount"
            }
        }

        with open(CONFIG_FILE, "w", encoding="utf-8") as file:
            json.dump(sample_config, file, ensure_ascii=False, indent=2)

        print(f"Sample config created: {CONFIG_FILE}")

    june_path = os.path.join(INPUT_DIR, "sample_sales_2026_06.xlsx")
    july_path = os.path.join(INPUT_DIR, "sample_sales_2026_07.xlsx")
    master_path = os.path.join(MASTER_DIR, "product_master.xlsx")

    if not os.path.exists(june_path):
        create_sample_sales_june_excel(june_path)
        print(f"Sample sales file created: {june_path}")

    if not os.path.exists(july_path):
        create_sample_sales_july_excel(july_path)
        print(f"Sample sales file created: {july_path}")

    if not os.path.exists(master_path):
        create_product_master_excel(master_path)
        print(f"Sample master file created: {master_path}")


def create_sample_sales_june_excel(file_path):
    """Create a sample June sales Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "sales_2026_06"

    headers = ["sales_id", "sales_date", "product_code", "quantity", "unit_price", "note"]

    rows = [
        ["S-202606-001", "2026/06/01", "P-1001", 2, 1800, "正常データ"],
        ["S-202606-002", "2026/06/02", "P-1002", 1, 3200, "正常データ"],
        ["S-202606-003", "2026/06/03", "P-1004", 1, 9800, "正常データ"],
        ["S-202606-004", "2026/06/04", "P-9999", 3, 1000, "マスタに存在しない商品コード"],
        ["S-202606-005", "2026/06/05", "P-1003", "abc", 4500, "数量が数値ではない"],
        ["S-202606-006", "not-date", "P-1001", 1, 1800, "日付が不正"],
    ]

    append_rows_to_worksheet(ws, headers, rows)
    wb.save(file_path)


def create_sample_sales_july_excel(file_path):
    """Create a sample July sales Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "sales_2026_07"

    headers = ["sales_id", "sales_date", "product_code", "quantity", "unit_price", "note"]

    rows = [
        ["S-202607-001", "2026/07/01", "P-1001", 5, 1750, "正常データ"],
        ["S-202607-002", "2026/07/02", "P-1003", 2, 4500, "正常データ"],
        ["S-202607-003", "2026/07/03", "P-1004", 2, 9600, "正常データ"],
        ["S-202607-004", "2026/07/04", "", 1, 3200, "商品コードが空欄"],
        ["S-202607-005", "2026/07/05", "P-1002", 2, "price?", "単価が数値ではない"],
        ["S-202607-006", "2026/07/06", "P-2000", 1, 12000, "マスタに存在しない商品コード"],
    ]

    append_rows_to_worksheet(ws, headers, rows)
    wb.save(file_path)


def create_product_master_excel(file_path):
    """Create a sample product master Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "product_master"

    headers = ["product_code", "product_name", "category"]

    rows = [
        ["P-1001", "ワイヤレスマウス", "PC周辺機器"],
        ["P-1002", "USBキーボード", "PC周辺機器"],
        ["P-1003", "Webカメラ", "PC周辺機器"],
        ["P-1004", "外付けSSD", "ストレージ"],
    ]

    append_rows_to_worksheet(ws, headers, rows)
    wb.save(file_path)


def append_rows_to_worksheet(worksheet, headers, rows):
    """Append headers and rows to a worksheet, then apply basic formatting."""
    worksheet.append(headers)

    for row in rows:
        worksheet.append(row)

    apply_basic_worksheet_format(worksheet)


def load_config():
    """Load config.json and normalize old/new setting names."""
    with open(CONFIG_FILE, "r", encoding="utf-8") as file:
        config = json.load(file)

    normalize_config_aliases(config)
    fill_config_defaults(config)
    return config


def normalize_config_aliases(config):
    """Support earlier config names while using the latest names internally."""
    if "sheet_names" not in config and "report_sheets" in config:
        config["sheet_names"] = config["report_sheets"]

    if "month_format" not in config and "sales_month_format" in config:
        config["month_format"] = config["sales_month_format"]

    if "sales_details_output_columns" not in config and "detail_output_columns" in config:
        config["sales_details_output_columns"] = config["detail_output_columns"]

    if "unmatched_products_output_columns" not in config and "unmatched_output_columns" in config:
        config["unmatched_products_output_columns"] = config["unmatched_output_columns"]

    if "validation_errors_output_columns" not in config and "unmatched_output_columns" in config:
        config["validation_errors_output_columns"] = config["unmatched_output_columns"]


def fill_config_defaults(config):
    """Fill optional config values with safe defaults."""
    sheet_names = DEFAULT_SHEET_NAMES.copy()
    user_sheet_names = config.get("sheet_names", {})

    if isinstance(user_sheet_names, dict):
        for key, value in user_sheet_names.items():
            if isinstance(value, str) and value.strip() != "":
                sheet_names[key] = value.strip()

    config["sheet_names"] = sheet_names
    config.setdefault("date_format", DEFAULT_DATE_FORMAT)
    config.setdefault("month_format", DEFAULT_MONTH_FORMAT)
    config.setdefault("amount_calculation", {"enabled": False})
    config.setdefault("sales_details_output_columns", DEFAULT_SALES_DETAILS_OUTPUT_COLUMNS)
    config.setdefault("monthly_sales_output_columns", DEFAULT_MONTHLY_SALES_OUTPUT_COLUMNS)
    config.setdefault("category_sales_output_columns", DEFAULT_CATEGORY_SALES_OUTPUT_COLUMNS)
    config.setdefault("product_sales_output_columns", DEFAULT_PRODUCT_SALES_OUTPUT_COLUMNS)
    config.setdefault("unmatched_products_output_columns", DEFAULT_ERROR_OUTPUT_COLUMNS)
    config.setdefault("validation_errors_output_columns", DEFAULT_ERROR_OUTPUT_COLUMNS)


def validate_config(config):
    """Validate config.json before processing files."""
    required_keys = [
        "sales_files",
        "master_file",
        "report_file",
        "sales_key",
        "master_key",
        "sales_required_columns",
        "master_required_columns",
        "master_value_columns",
        "sales_numeric_columns",
        "sales_date_columns",
        "sales_details_output_columns",
        "monthly_sales_output_columns",
        "category_sales_output_columns",
        "product_sales_output_columns",
        "unmatched_products_output_columns",
        "validation_errors_output_columns",
    ]

    missing_keys = [key for key in required_keys if key not in config]

    if missing_keys:
        raise ValueError(f"Missing config keys: {', '.join(missing_keys)}")

    validate_string_list(config, "sales_files", allow_empty=False)

    for key in ["master_file", "report_file", "sales_key", "master_key", "date_format", "month_format"]:
        validate_string_value(config, key)

    for key in [
        "sales_required_columns",
        "master_required_columns",
        "master_value_columns",
        "sales_numeric_columns",
        "sales_date_columns",
        "sales_details_output_columns",
        "monthly_sales_output_columns",
        "category_sales_output_columns",
        "product_sales_output_columns",
        "unmatched_products_output_columns",
        "validation_errors_output_columns",
    ]:
        validate_string_list(config, key, allow_empty=True)

    for sales_file in config["sales_files"]:
        validate_file_extension(sales_file, SUPPORTED_INPUT_EXTENSIONS, "sales_files")

    validate_file_extension(config["master_file"], SUPPORTED_INPUT_EXTENSIONS, "master_file")
    validate_file_extension(config["report_file"], [EXCEL_EXTENSION], "report_file")

    if config["sales_key"] not in config["sales_required_columns"]:
        raise ValueError("sales_key must be included in sales_required_columns.")

    if config["master_key"] not in config["master_required_columns"]:
        raise ValueError("master_key must be included in master_required_columns.")

    for column in config["master_value_columns"]:
        if column == config["master_key"]:
            raise ValueError("master_value_columns must not include master_key.")

    validate_amount_calculation(config)
    validate_sheet_names(config)


def validate_amount_calculation(config):
    """Validate amount_calculation settings."""
    amount_config = config.get("amount_calculation", {"enabled": False})

    if not isinstance(amount_config, dict):
        raise ValueError("amount_calculation must be an object.")

    if not amount_config.get("enabled", False):
        return

    required_keys = ["quantity_column", "price_column", "output_column"]
    missing_keys = [key for key in required_keys if key not in amount_config]

    if missing_keys:
        raise ValueError(f"Missing amount_calculation keys: {', '.join(missing_keys)}")

    for key in required_keys:
        if not isinstance(amount_config[key], str) or amount_config[key].strip() == "":
            raise ValueError(f"amount_calculation.{key} must be a non-empty string.")

    quantity_column = amount_config["quantity_column"]
    price_column = amount_config["price_column"]
    output_column = amount_config["output_column"]

    if quantity_column not in config["sales_numeric_columns"]:
        raise ValueError("amount_calculation.quantity_column must be included in sales_numeric_columns.")

    if price_column not in config["sales_numeric_columns"]:
        raise ValueError("amount_calculation.price_column must be included in sales_numeric_columns.")

    if output_column not in config["sales_details_output_columns"]:
        raise ValueError("amount_calculation.output_column must be included in sales_details_output_columns.")


def validate_sheet_names(config):
    """Validate sheet name settings."""
    sheet_names = config.get("sheet_names", {})

    if not isinstance(sheet_names, dict):
        raise ValueError("sheet_names must be an object.")

    invalid_values = [
        value for value in sheet_names.values()
        if not isinstance(value, str) or value.strip() == ""
    ]

    if invalid_values:
        raise ValueError("sheet_names must contain only non-empty strings.")

    duplicated_names = find_duplicates(list(sheet_names.values()))

    if duplicated_names:
        raise ValueError("sheet_names must not contain duplicated names: " + ", ".join(duplicated_names))


def validate_string_value(config, key):
    """Validate that config[key] is a non-empty string."""
    if not isinstance(config[key], str) or config[key].strip() == "":
        raise ValueError(f"{key} must be a non-empty string.")


def validate_string_list(config, key, allow_empty):
    """Validate that config[key] is a list of non-empty strings."""
    if not isinstance(config[key], list):
        raise ValueError(f"{key} must be a list.")

    if not allow_empty and len(config[key]) == 0:
        raise ValueError(f"{key} must not be empty.")

    invalid_values = [
        value for value in config[key]
        if not isinstance(value, str) or value.strip() == ""
    ]

    if invalid_values:
        raise ValueError(f"{key} must contain only non-empty strings.")


def validate_file_extension(file_name, allowed_extensions, config_key):
    """Validate a file extension against allowed extensions."""
    extension = os.path.splitext(file_name)[1].lower()

    if extension not in allowed_extensions:
        allowed_text = ", ".join(allowed_extensions)
        raise ValueError(f"{config_key} must be one of these types: {allowed_text}")


def find_duplicates(values):
    """Return duplicated values while keeping deterministic order."""
    seen_values = set()
    duplicated_values = []

    for value in values:
        if value in seen_values and value not in duplicated_values:
            duplicated_values.append(value)
        seen_values.add(value)

    return duplicated_values


def make_sales_path(sales_file):
    """Return a sales file path under input directory."""
    return os.path.join(INPUT_DIR, sales_file)


def make_master_path(config):
    """Return the master file path under master directory."""
    return os.path.join(MASTER_DIR, config["master_file"])


def make_report_path(config):
    """Return the report file path under output directory."""
    return os.path.join(OUTPUT_DIR, config["report_file"])


def read_table_file(file_path):
    """Read a CSV or Excel file as a DataFrame."""
    extension = os.path.splitext(file_path)[1].lower()

    if extension == CSV_EXTENSION:
        df = read_csv_with_fallback_encodings(file_path)
    elif extension == EXCEL_EXTENSION:
        df = pd.read_excel(file_path, dtype=object)
    else:
        raise ValueError(f"Unsupported file type: {file_path}")

    return normalize_column_names(df, file_path)


def read_csv_with_fallback_encodings(file_path):
    """Read a CSV file by trying common Japanese encodings."""
    last_error = None

    for encoding in DEFAULT_CSV_READ_ENCODINGS:
        try:
            return pd.read_csv(file_path, dtype=object, encoding=encoding)
        except UnicodeDecodeError as error:
            last_error = error

    raise UnicodeDecodeError(
        last_error.encoding,
        last_error.object,
        last_error.start,
        last_error.end,
        f"Failed to read CSV with encodings: {', '.join(DEFAULT_CSV_READ_ENCODINGS)}"
    )


def normalize_column_names(df, file_path):
    """Strip column names and reject duplicated column names."""
    df = df.copy()
    df.columns = [str(column).strip() for column in df.columns]

    duplicated_columns = find_duplicates(list(df.columns))

    if duplicated_columns:
        raise ValueError(
            f"Duplicated columns found in {file_path}: "
            f"{', '.join(duplicated_columns)}"
        )

    return df


def read_sales_files(config):
    """Read all configured sales files and combine them into one DataFrame."""
    sales_frames = []

    for sales_file in config["sales_files"]:
        sales_path = make_sales_path(sales_file)

        if not os.path.exists(sales_path):
            raise FileNotFoundError(f"Sales file not found: {sales_path}")

        sales_df = read_table_file(sales_path)
        check_required_columns(sales_df, config["sales_required_columns"], f"sales file: {sales_file}")
        prepared_sales_df = prepare_sales_dataframe(sales_df, sales_file)
        sales_frames.append(prepared_sales_df)

        print(f"Sales file loaded: {sales_path} ({len(prepared_sales_df)} rows)")

    if not sales_frames:
        return pd.DataFrame()

    return pd.concat(sales_frames, ignore_index=True)


def check_required_columns(df, required_columns, file_label):
    """Ensure all required columns exist in a DataFrame."""
    missing_columns = [
        column for column in required_columns
        if column not in df.columns
    ]

    if missing_columns:
        raise ValueError(f"Missing columns in {file_label}: {', '.join(missing_columns)}")


def normalize_dataframe_values(df):
    """Normalize all visible cell values in a DataFrame."""
    normalized_df = df.copy()
    return normalized_df.apply(lambda column: column.map(normalize_general_value))


def prepare_sales_dataframe(sales_df, source_file):
    """Normalize sales values and add source file and original row numbers."""
    prepared_df = normalize_dataframe_values(sales_df)
    prepared_df[SOURCE_FILE_COLUMN] = source_file
    prepared_df[ROW_NUMBER_COLUMN] = list(range(2, len(prepared_df) + 2))
    return prepared_df


def prepare_master_dataframe(master_df):
    """Normalize master values and add original row numbers."""
    prepared_df = normalize_dataframe_values(master_df)
    prepared_df[INTERNAL_MASTER_ROW_NUMBER] = list(range(2, len(prepared_df) + 2))
    return prepared_df


def validate_and_prepare_master(master_df, config):
    """Validate master data and return columns used for matching and enrichment."""
    master_key = config["master_key"]
    required_columns = config["master_required_columns"]
    error_messages = []

    for _, row in master_df.iterrows():
        row_number = row[INTERNAL_MASTER_ROW_NUMBER]

        for column in required_columns:
            if is_blank(row.get(column, "")):
                error_messages.append(f"row {row_number}: {column} is required")

    blank_key_rows = [
        str(row[INTERNAL_MASTER_ROW_NUMBER])
        for _, row in master_df.iterrows()
        if is_blank(row.get(master_key, ""))
    ]

    if blank_key_rows:
        error_messages.append(f"master_key has blank values at rows: {', '.join(blank_key_rows)}")

    duplicated_key_rows = find_duplicate_master_key_rows(master_df, master_key)

    if duplicated_key_rows:
        error_messages.append("master_key has duplicated values: " + "; ".join(duplicated_key_rows))

    if error_messages:
        raise ValueError("Master data is invalid. " + " | ".join(error_messages))

    keep_columns = [master_key] + config["master_value_columns"]
    return master_df[keep_columns].copy()


def find_duplicate_master_key_rows(master_df, master_key):
    """Return readable duplicate-key information for master data."""
    duplicated_info = []
    non_blank_df = master_df[~master_df[master_key].map(is_blank)].copy()
    duplicated_mask = non_blank_df.duplicated(subset=[master_key], keep=False)
    duplicated_df = non_blank_df[duplicated_mask]

    for key_value, group_df in duplicated_df.groupby(master_key, sort=True):
        row_numbers = [str(value) for value in group_df[INTERNAL_MASTER_ROW_NUMBER].tolist()]
        duplicated_info.append(f"{key_value} at rows {', '.join(row_numbers)}")

    return duplicated_info


def split_validation_errors(sales_df, config):
    """Split sales rows into valid rows and validation error rows."""
    valid_rows = []
    error_rows = []

    for _, row in sales_df.iterrows():
        row_dict = row.to_dict()
        error_messages = []

        check_required_values(row_dict, config["sales_required_columns"], error_messages)
        normalize_numeric_values(row_dict, config["sales_numeric_columns"], error_messages)
        normalize_date_values(
            row_dict,
            config["sales_date_columns"],
            config.get("date_format", DEFAULT_DATE_FORMAT),
            error_messages,
        )

        if error_messages:
            error_rows.append(
                create_error_record(
                    row_dict,
                    config["validation_errors_output_columns"],
                    error_type="validation_error",
                    error_message="; ".join(error_messages),
                )
            )
        else:
            valid_rows.append(row_dict)

    valid_df = pd.DataFrame(valid_rows)
    error_df = pd.DataFrame(error_rows)

    valid_df = ensure_columns(valid_df, list(sales_df.columns))
    error_df = select_output_columns(error_df, config["validation_errors_output_columns"])

    return valid_df, error_df


def check_required_values(row_dict, required_columns, error_messages):
    """Add an error when a required value is blank."""
    for column in required_columns:
        if is_blank(row_dict.get(column, "")):
            error_messages.append(f"{column} is required")


def normalize_numeric_values(row_dict, numeric_columns, error_messages):
    """Validate numeric columns and keep valid values as numbers."""
    for column in numeric_columns:
        value = row_dict.get(column, "")

        if is_blank(value):
            row_dict[column] = ""
            continue

        if not can_convert_to_finite_number(value):
            error_messages.append(f"{column} must be numeric")
            continue

        row_dict[column] = convert_to_clean_number(value)


def normalize_date_values(row_dict, date_columns, date_format, error_messages):
    """Validate date columns and format them for output."""
    for column in date_columns:
        value = row_dict.get(column, "")

        if is_blank(value):
            row_dict[column] = ""
            continue

        try:
            parsed_date = pd.to_datetime(value, errors="raise")
        except Exception:
            error_messages.append(f"{column} must be a valid date")
            continue

        row_dict[column] = parsed_date.strftime(date_format)


def normalize_general_value(value):
    """Normalize a general cell value to a safe string value."""
    if is_blank(value):
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    return str(value).strip()


def is_blank(value):
    """Return True when value should be treated as blank."""
    if value is None:
        return True

    try:
        if pd.isna(value):
            return True
    except TypeError:
        pass

    if isinstance(value, str) and value.strip() == "":
        return True

    return False


def can_convert_to_finite_number(value):
    """Return True when value can be converted to a finite number."""
    try:
        numeric_value = pd.to_numeric(value)
        return math.isfinite(float(numeric_value))
    except Exception:
        return False


def convert_to_clean_number(value):
    """Convert a numeric-like value to int when possible, otherwise float."""
    numeric_value = float(pd.to_numeric(value))

    if numeric_value.is_integer():
        return int(numeric_value)

    return numeric_value


def match_with_master(valid_sales_df, master_df, config):
    """Match valid sales rows with master rows by configured keys."""
    sales_key = config["sales_key"]
    master_key = config["master_key"]

    if len(valid_sales_df) == 0:
        columns = list(valid_sales_df.columns) + list(master_df.columns) + [MERGE_STATUS_COLUMN]
        return pd.DataFrame(columns=list(dict.fromkeys(columns)))

    if sales_key == master_key:
        return pd.merge(
            valid_sales_df,
            master_df,
            how="left",
            on=sales_key,
            indicator=True,
            validate="many_to_one",
        )

    return pd.merge(
        valid_sales_df,
        master_df,
        how="left",
        left_on=sales_key,
        right_on=master_key,
        indicator=True,
        validate="many_to_one",
    )


def build_sales_details_df(merged_df, config):
    """Build detail rows successfully matched with master data."""
    if len(merged_df) == 0:
        return pd.DataFrame(columns=config["sales_details_output_columns"])

    matched_df = merged_df[merged_df[MERGE_STATUS_COLUMN] == "both"].copy()

    if len(matched_df) == 0:
        return pd.DataFrame(columns=config["sales_details_output_columns"])

    add_sales_amount_column(matched_df, config)
    add_sales_month_column(matched_df, config)

    details_df = select_output_columns(matched_df, config["sales_details_output_columns"])
    return sort_by_source_and_row(details_df)


def add_sales_amount_column(df, config):
    """Add sales amount column when amount calculation is enabled."""
    amount_config = config.get("amount_calculation", {"enabled": False})

    if not amount_config.get("enabled", False):
        return

    quantity_column = amount_config["quantity_column"]
    price_column = amount_config["price_column"]
    output_column = amount_config["output_column"]

    df[output_column] = pd.to_numeric(df[quantity_column]) * pd.to_numeric(df[price_column])


def add_sales_month_column(df, config):
    """Add sales_month column based on the first configured sales date column."""
    if not config["sales_date_columns"]:
        return

    sales_date_column = config["sales_date_columns"][0]

    if sales_date_column not in df.columns:
        return

    month_format = config.get("month_format", DEFAULT_MONTH_FORMAT)
    df["sales_month"] = pd.to_datetime(df[sales_date_column]).dt.strftime(month_format)


def build_unmatched_products_df(merged_df, config):
    """Build rows whose product codes were not found in master data."""
    if len(merged_df) == 0:
        return pd.DataFrame(columns=config["unmatched_products_output_columns"])

    master_not_found_df = merged_df[merged_df[MERGE_STATUS_COLUMN] == "left_only"].copy()

    if len(master_not_found_df) == 0:
        return pd.DataFrame(columns=config["unmatched_products_output_columns"])

    records = []

    for _, row in master_not_found_df.iterrows():
        records.append(
            create_error_record(
                row.to_dict(),
                config["unmatched_products_output_columns"],
                error_type="master_not_found",
                error_message=f"{config['sales_key']} not found in master",
            )
        )

    unmatched_df = pd.DataFrame(records)
    unmatched_df = select_output_columns(unmatched_df, config["unmatched_products_output_columns"])
    return sort_by_source_and_row(unmatched_df)


def create_error_record(row_dict, output_columns, error_type, error_message):
    """Create one normalized error output record."""
    record = {}

    for column in output_columns:
        if column == ERROR_TYPE_COLUMN:
            record[column] = error_type
        elif column == ERROR_MESSAGE_COLUMN:
            record[column] = error_message
        else:
            record[column] = row_dict.get(column, "")

    return record


def build_monthly_sales_df(details_df, config):
    """Build monthly sales summary."""
    base_columns = ["sales_month", "sales_count", "total_quantity", "total_sales"]

    if len(details_df) == 0 or "sales_month" not in details_df.columns:
        return pd.DataFrame(columns=config["monthly_sales_output_columns"])

    amount_column = get_amount_output_column(config)
    quantity_column = get_quantity_column(config)

    monthly_df = (
        details_df
        .groupby("sales_month", dropna=False)
        .agg(
            sales_count=(config["sales_required_columns"][0], "count"),
            total_quantity=(quantity_column, "sum"),
            total_sales=(amount_column, "sum"),
        )
        .reset_index()
        .sort_values(by="sales_month", kind="stable")
    )

    monthly_df = select_output_columns(monthly_df, base_columns)
    return select_output_columns(monthly_df, config["monthly_sales_output_columns"])


def build_category_sales_df(details_df, config):
    """Build category sales summary."""
    base_columns = ["category", "sales_count", "total_quantity", "total_sales"]

    if len(details_df) == 0 or "category" not in details_df.columns:
        return pd.DataFrame(columns=config["category_sales_output_columns"])

    amount_column = get_amount_output_column(config)
    quantity_column = get_quantity_column(config)

    category_df = (
        details_df
        .groupby("category", dropna=False)
        .agg(
            sales_count=(config["sales_required_columns"][0], "count"),
            total_quantity=(quantity_column, "sum"),
            total_sales=(amount_column, "sum"),
        )
        .reset_index()
        .sort_values(by="total_sales", ascending=False, kind="stable")
    )

    category_df = select_output_columns(category_df, base_columns)
    return select_output_columns(category_df, config["category_sales_output_columns"])


def build_product_sales_df(details_df, config):
    """Build product sales ranking."""
    base_columns = [
        "product_code",
        "product_name",
        "category",
        "sales_count",
        "total_quantity",
        "total_sales",
    ]

    group_columns = ["product_code", "product_name", "category"]

    if len(details_df) == 0 or any(column not in details_df.columns for column in group_columns):
        return pd.DataFrame(columns=config["product_sales_output_columns"])

    amount_column = get_amount_output_column(config)
    quantity_column = get_quantity_column(config)

    product_df = (
        details_df
        .groupby(group_columns, dropna=False)
        .agg(
            sales_count=(config["sales_required_columns"][0], "count"),
            total_quantity=(quantity_column, "sum"),
            total_sales=(amount_column, "sum"),
        )
        .reset_index()
        .sort_values(by="total_sales", ascending=False, kind="stable")
    )

    product_df = select_output_columns(product_df, base_columns)
    return select_output_columns(product_df, config["product_sales_output_columns"])


def build_summary_df(
    config,
    total_sales_rows,
    master_rows,
    valid_sales_df,
    details_df,
    unmatched_products_df,
    validation_errors_df,
    started_at,
    finished_at,
):
    """Build report summary DataFrame."""
    total_quantity = 0
    total_sales = 0
    period_start = ""
    period_end = ""

    quantity_column = get_quantity_column(config)
    amount_column = get_amount_output_column(config)

    if len(details_df) > 0:
        if quantity_column in details_df.columns:
            total_quantity = details_df[quantity_column].sum()

        if amount_column in details_df.columns:
            total_sales = details_df[amount_column].sum()

        date_column = config["sales_date_columns"][0] if config["sales_date_columns"] else ""

        if date_column in details_df.columns:
            parsed_dates = pd.to_datetime(details_df[date_column], errors="coerce").dropna()

            if len(parsed_dates) > 0:
                period_start = parsed_dates.min().strftime(config.get("date_format", DEFAULT_DATE_FORMAT))
                period_end = parsed_dates.max().strftime(config.get("date_format", DEFAULT_DATE_FORMAT))

    summary_records = [
        {"item": "sales_files", "value": ", ".join(config["sales_files"])},
        {"item": "master_file", "value": config["master_file"]},
        {"item": "report_file", "value": config["report_file"]},
        {"item": "total_files", "value": len(config["sales_files"])},
        {"item": "total_sales_rows", "value": total_sales_rows},
        {"item": "master_rows", "value": master_rows},
        {"item": "valid_sales_rows", "value": len(valid_sales_df)},
        {"item": "sales_detail_rows", "value": len(details_df)},
        {"item": "validation_error_rows", "value": len(validation_errors_df)},
        {"item": "unmatched_product_rows", "value": len(unmatched_products_df)},
        {"item": "total_quantity", "value": total_quantity},
        {"item": "total_sales", "value": total_sales},
        {"item": "period_start", "value": period_start},
        {"item": "period_end", "value": period_end},
        {"item": "sales_key", "value": config["sales_key"]},
        {"item": "master_key", "value": config["master_key"]},
        {"item": "started_at", "value": started_at.strftime("%Y-%m-%d %H:%M:%S")},
        {"item": "finished_at", "value": finished_at.strftime("%Y-%m-%d %H:%M:%S")},
    ]

    return pd.DataFrame(summary_records, columns=["item", "value"])


def get_quantity_column(config):
    """Return configured quantity column used for aggregation."""
    return config.get("amount_calculation", {}).get("quantity_column", "quantity")


def get_amount_output_column(config):
    """Return configured amount output column used for aggregation."""
    return config.get("amount_calculation", {}).get("output_column", "sales_amount")


def ensure_columns(df, output_columns):
    """Ensure a DataFrame has all requested columns."""
    output_df = df.copy()

    for column in output_columns:
        if column not in output_df.columns:
            output_df[column] = ""

    return output_df[output_columns].copy()


def select_output_columns(df, output_columns):
    """Select output columns while creating missing columns as blanks."""
    if len(df) == 0:
        return pd.DataFrame(columns=output_columns)

    return ensure_columns(df, output_columns)


def sort_by_source_and_row(df):
    """Sort rows by source file and row number when possible."""
    sort_columns = [
        column for column in [SOURCE_FILE_COLUMN, ROW_NUMBER_COLUMN]
        if column in df.columns
    ]

    if sort_columns and len(df) > 0:
        return df.sort_values(by=sort_columns, kind="stable").reset_index(drop=True)

    return df.reset_index(drop=True)


def save_report(
    summary_df,
    details_df,
    monthly_sales_df,
    category_sales_df,
    product_sales_df,
    unmatched_products_df,
    validation_errors_df,
    config,
):
    """Save the final sales summary report as a multi-sheet Excel file."""
    report_path = make_report_path(config)
    sheet_names = config["sheet_names"]

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name=sheet_names["summary"], index=False)
        details_df.to_excel(writer, sheet_name=sheet_names["sales_details"], index=False)
        monthly_sales_df.to_excel(writer, sheet_name=sheet_names["monthly_sales"], index=False)
        category_sales_df.to_excel(writer, sheet_name=sheet_names["category_sales"], index=False)
        product_sales_df.to_excel(writer, sheet_name=sheet_names["product_sales"], index=False)
        unmatched_products_df.to_excel(writer, sheet_name=sheet_names["unmatched_products"], index=False)
        validation_errors_df.to_excel(writer, sheet_name=sheet_names["validation_errors"], index=False)

        for worksheet in writer.book.worksheets:
            apply_basic_worksheet_format(worksheet)

    print(f"Sales summary report saved: {report_path}")


def apply_basic_worksheet_format(worksheet):
    """Apply basic formatting to an openpyxl worksheet."""
    format_header_row(worksheet)
    adjust_column_width(worksheet)
    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions


def format_header_row(worksheet):
    """Apply a simple header style to the first row."""
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font


def adjust_column_width(worksheet):
    """Adjust column widths based on cell values."""
    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def main():
    started_at = datetime.now()

    create_sample_files()

    config = load_config()
    validate_config(config)

    master_path = make_master_path(config)

    if not os.path.exists(master_path):
        raise FileNotFoundError(f"Master file not found: {master_path}")

    print(f"Sales files: {', '.join(config['sales_files'])}")
    print(f"Master file: {master_path}")
    print(f"Report file: {make_report_path(config)}")

    sales_df = read_sales_files(config)
    master_df = read_table_file(master_path)

    check_required_columns(master_df, config["master_required_columns"], "master file")
    check_required_columns(
        master_df,
        [config["master_key"]] + config["master_value_columns"],
        "master file",
    )

    total_sales_rows = len(sales_df)
    master_rows = len(master_df)

    print(f"Total sales rows found: {total_sales_rows}")
    print(f"Master rows found: {master_rows}")

    prepared_master_df = prepare_master_dataframe(master_df)
    prepared_master_df = validate_and_prepare_master(prepared_master_df, config)

    valid_sales_df, validation_errors_df = split_validation_errors(sales_df, config)
    merged_df = match_with_master(valid_sales_df, prepared_master_df, config)

    details_df = build_sales_details_df(merged_df, config)
    unmatched_products_df = build_unmatched_products_df(merged_df, config)
    validation_errors_df = sort_by_source_and_row(validation_errors_df)

    monthly_sales_df = build_monthly_sales_df(details_df, config)
    category_sales_df = build_category_sales_df(details_df, config)
    product_sales_df = build_product_sales_df(details_df, config)

    finished_at = datetime.now()

    summary_df = build_summary_df(
        config,
        total_sales_rows,
        master_rows,
        valid_sales_df,
        details_df,
        unmatched_products_df,
        validation_errors_df,
        started_at,
        finished_at,
    )

    save_report(
        summary_df,
        details_df,
        monthly_sales_df,
        category_sales_df,
        product_sales_df,
        unmatched_products_df,
        validation_errors_df,
        config,
    )

    print("Sales summary reporting finished.")
    print(f"Valid sales rows: {len(valid_sales_df)}")
    print(f"Sales detail rows: {len(details_df)}")
    print(f"Validation error rows: {len(validation_errors_df)}")
    print(f"Unmatched product rows: {len(unmatched_products_df)}")
    print(f"Output folder: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
